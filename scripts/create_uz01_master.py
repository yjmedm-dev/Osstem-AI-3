# -*- coding: utf-8 -*-
"""UZ01 계정과목 마스터 양식 엑셀 생성 스크립트
실행: python scripts/create_uz01_master.py
출력: data/reference/UZ01_계정마스터_양식.xlsx
"""
import pathlib
import sys
import yaml
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# ── 1. accounts_master.yaml → standard_code 조회 테이블 ──────────────────────
with open(ROOT / "config/accounts_master.yaml", encoding="utf-8") as f:
    yaml_data = yaml.safe_load(f)

std_lookup: dict[str, dict] = {}
for acc in yaml_data.get("accounts", []):
    std_lookup[acc["code"]] = {"name": acc["name"], "type": acc["type"]}

# ── 2. UZ01 데이터 (local_code, local_name, standard_code) ───────────────────
# local_code : 1C 우즈베키스탄 계정코드 (숫자) 또는 러시아어 비용 항목명
UZ01_ROWS = [
    # ── 비유동자산 ──────────────────────────────────────────────
    ("0120.2",  "구조물",                       "FP01-02-02-0160"),
    ("0130.1",  "동력기계",                     "FP01-02-02-0160"),
    ("0130.2",  "작업기계",                     "FP01-02-02-0160"),
    ("0130.4",  "기타기계",                     "FP01-02-02-0160"),
    ("0140",    "가구·사무기기",                "FP01-02-02-0160"),
    ("0150",    "컴퓨터",                       "FP01-02-02-0160"),
    ("0160",    "차량",                         "FP01-02-02-0110"),
    ("0190",    "기타유형자산",                 "FP01-02-02-0160"),
    # 감가상각누계
    ("0220.2",  "구조물 감가상각",              "FP01-02-02-0170"),
    ("0230.1",  "동력기계 감가상각",            "FP01-02-02-0170"),
    ("0230.2",  "작업기계 감가상각",            "FP01-02-02-0170"),
    ("0230.4",  "기타기계 감가상각",            "FP01-02-02-0170"),
    ("0240",    "가구 감가상각",                "FP01-02-02-0170"),
    ("0250",    "컴퓨터 감가상각",              "FP01-02-02-0170"),
    ("0260",    "차량 감가상각",                "FP01-02-02-0120"),
    ("0290",    "기타유형자산 감가상각",        "FP01-02-02-0170"),
    # 무형자산
    ("0430",    "소프트웨어",                   "FP01-02-03-0050"),
    ("0530",    "소프트웨어 상각",              "FP01-02-03-0050"),
    # ── 재고자산 ────────────────────────────────────────────────
    ("0820.1",  "유형자산 구성품(1)",           "FP01-01-02-0110"),
    ("0820.2",  "유형자산 구성품(2)",           "FP01-01-02-0110"),
    ("0830",    "유형자산 취득중",              "FP01-01-02-0110"),
    ("1010",    "원재료",                       "FP01-01-02-0110"),
    ("1030",    "연료",                         "FP01-01-02-0110"),
    ("1040",    "예비부품",                     "FP01-01-02-0110"),
    ("1080",    "비품 재료",                    "FP01-01-02-0110"),
    ("1090",    "기타 재료",                    "FP01-01-02-0110"),
    ("1510",    "재료 취득·구매",               "FP01-01-02-0110"),
    ("2910",    "창고 상품",                    "FP01-01-02-0010-02"),
    ("2920.2",  "상품(2)",                      "FP01-01-02-0010-02"),
    ("2920.3",  "상품(3)",                      "FP01-01-02-0010-02"),
    ("2970.1",  "운송중 상품(국내)",            "FP01-01-02-0120"),
    ("2970.2",  "운송중 상품(국외)",            "FP01-01-02-0120"),
    ("2980.2",  "기타상품(2)",                  "FP01-01-02-0010-02"),
    ("2980.3",  "기타상품(3)",                  "FP01-01-02-0010-02"),
    # ── 당좌자산 ────────────────────────────────────────────────
    ("3190",    "선급비용",                     "FP01-01-01-0170"),
    ("4010",    "매출채권",                     "FP01-01-01-0070"),
    ("4015",    "매출채권(외화)",               "FP01-01-01-0070"),
    ("4090",    "소매 매출채권",                "FP01-01-01-0070"),
    ("4210",    "급여 선급",                    "FP01-01-01-0170"),
    ("4220",    "출장 선급",                    "FP01-01-01-0170"),
    ("4230",    "일반관리 선급",                "FP01-01-01-0170"),
    ("4310",    "선급금",                       "FP01-01-01-0170"),
    ("4315",    "선급금(외화)",                 "FP01-01-01-0170"),
    ("4410.1",  "선급부가세(유형자산)",         "FP01-01-01-0210"),
    ("4410.2",  "선급부가세(상품)",             "FP01-01-01-0210"),
    ("4410.3",  "선급부가세(서비스)",           "FP01-01-01-0210"),
    ("4410.4",  "선급부가세(기타1)",            "FP01-01-01-0210"),
    ("4410.5",  "선급부가세(기타2)",            "FP01-01-01-0210"),
    ("4410.10", "선급부가세(기타3)",            "FP01-01-01-0210"),
    ("4420.1",  "선급소득세",                   "FP01-01-01-0190"),
    ("4420.3",  "선급개인소득세",               "FP01-01-01-0190"),
    ("4430",    "선납법인세",                   "FP01-01-01-0200"),
    ("4490",    "기타세금 선납",                "FP01-01-01-0210"),
    ("4510.1",  "사회보험 선납(1)",             "FP01-01-01-0170"),
    ("4510.2",  "사회보험 선납(2)",             "FP01-01-01-0170"),
    ("4530.1",  "INPS 단기대여",               "FP01-01-01-0140"),
    ("4720",    "대여금",                       "FP01-01-01-0170"),
    ("4820",    "임대 보증금",                  "FP01-02-04-0030"),
    ("4890",    "기타 채권",                    "FP01-02-04-0030"),
    # 현금 및 현금성자산
    ("5010",    "현금",                         "FP01-01-01-0010"),
    ("5050",    "소매 현금",                    "FP01-01-01-0010"),
    ("5110",    "당좌예금",                     "FP01-01-01-0010"),
    ("5210",    "외화예금",                     "FP01-01-01-0010"),
    ("5530",    "특수예금(1)",                  "FP01-01-01-0010"),
    ("5535",    "특수예금(2)",                  "FP01-01-01-0010"),
    ("5710",    "이체중 자금(1)",               "FP01-01-01-0100"),
    ("5711",    "이체중 자금(2)",               "FP01-01-01-0100"),
    ("5720",    "카드매출 미수",                "FP01-01-01-0100"),
    ("5730",    "외화매입 미수",                "FP01-01-01-0100"),
    # ── 유동부채 ────────────────────────────────────────────────
    ("6010",    "매입채무",                     "FP02-01-01-0010"),
    ("6015",    "매입채무(외화)",               "FP02-01-01-0010"),
    ("6310",    "선수금",                       "FP02-01-01-0120"),
    ("6315",    "선수금(외화)",                 "FP02-01-01-0120"),
    ("6410.1",  "부가세예수금(1)",              "FP02-01-01-0180"),
    ("6410.2",  "부가세예수금(2)",              "FP02-01-01-0180"),
    ("6420.1",  "미지급 소득세",               "FP02-01-01-0150"),
    ("6420.3",  "미지급 개인소득세",           "FP02-01-01-0150"),
    ("6430",    "미지급 법인세",               "FP02-01-01-0150"),
    ("6490",    "기타세금 미지급",             "FP02-01-01-0110"),
    ("6510.1",  "사회보험 미지급",             "FP02-01-01-0110"),
    ("6530.1",  "INPS 미지급",                "FP02-01-01-0110"),
    ("6710",    "미지급 급여",                 "FP02-01-01-0150"),
    ("6910",    "임차료 미지급",               "FP02-01-01-0110"),
    ("6970",    "가불금",                      "FP02-01-01-0110"),
    ("6975",    "기타채무(1)",                 "FP02-01-01-0110"),
    ("6990",    "기타채무",                    "FP02-01-01-0110"),
    ("6990.1",  "수입채무",                    "FP02-01-01-0010"),
    # ── 자본 ────────────────────────────────────────────────────
    ("8330",    "출자금",                      "FP03-01-01-0010"),
    ("8420",    "설립시 환차",                 "FP03-01-01-0010"),
    ("8510.1",  "자산재평가",                  "FP03-03-01-0070"),
    ("8530.3",  "무상수령",                    "PL06-00-0280"),
    ("8710",    "이월잉여금",                  "FP03-05-01-0060"),
    # ── 손익 — 매출·매출원가 ────────────────────────────────────
    ("9020.1",  "상품매출",                    "PL01-01-0020"),
    ("9040.1",  "매출반품",                    "PL01-01-0020"),
    ("9120.1",  "매출원가",                    "PL02-01-0020"),
    # ── 손익 — 영업외 ───────────────────────────────────────────
    ("9310",    "유형자산처분이익",            "PL06-00-0280"),
    ("9380",    "무상재정지원",                "PL06-00-0280"),
    ("9390",    "기타영업수익",                "PL06-00-0280"),
    ("9540",    "외화환산이익",                "PL06-00-0050"),
    ("9541",    "선급금 환산이익",             "PL06-00-0050"),
    ("9690",    "기타금융비용",                "PL07-00-0260"),
    ("9810",    "법인세",                      "PL09-00-0010"),
    ("9820",    "기타의무납부세",              "PL09-00-0010"),
    ("9910",    "당기손익",                    "FP03-05-01-0070"),
    # ── PL — 판관비 (러시아어 비용 항목명 직접 매핑) ─────────────
    ("Amortizatsiya 자산",                     "감가상각비(유형자산)",    "PL04-00-0180"),
    ("Amortizatsiya 무형",                     "감가상각비(무형자산)",    "PL04-00-0200"),
    ("Arenda nezhilogo pomeshcheniya zdaniya", "임차료(건물)",            "PL04-00-0140"),
    ("Arenda ofisa (na 5 let)",                "임차료(사무실 5년)",      "PL04-00-0140"),
    ("Bonusy",                                 "상여금",                  "PL04-00-0020"),
    ("GSM",                                    "차량유지비(유류)",         "PL04-00-0170"),
    ("Donachislennyy NDS",                     "세금과공과(부가세추가)",  "PL04-00-0230"),
    ("Internet",                               "통신비(인터넷)",          "PL04-00-0120"),
    ("Komandirovochnye raskhody",              "여비교통비(출장)",         "PL04-00-0070"),
    ("Kureryerskie uslugi (dostavka)",         "운반비(택배)",             "PL04-00-0110"),
    ("Kursovye raznitsy",                      "외화환산손실",             "PL07-00-0040"),
    ("Nalog s oborota s samozanyatogo litsa",  "복리후생비(세금)",         "PL04-00-0050"),
    ("Oplata bolnichnykh listov",              "복리후생비(병가)",         "PL04-00-0050"),
    ("Oplata truda",                           "급여",                    "PL04-00-0010"),
    ("Otkloneniya kursa prodazhi inostrannoy valyuty", "외환차손",        "PL07-00-0030"),
    ("Periodicheskie elektronnye izdaniya",    "수도광열비(정기간행물)",  "PL04-00-0130"),
    ("Povyshenie kvalifikatsii sotrudnikov",   "복리후생비(교육훈련)",    "PL04-00-0050"),
    ("Poisk personala",                        "지급수수료(채용)",         "PL04-00-0240"),
    ("Provedenie Konferentsiy",                "광고선전비(컨퍼런스)",    "PL04-00-0210"),
    ("Razmeshchenie reklamy",                  "광고선전비(광고게재)",    "PL04-00-0210"),
    ("Raskhody na marketing i reklamu",        "광고선전비(마케팅)",       "PL04-00-0210"),
    ("Raskhody na obrabotku",                  "지급수수료(처리)",         "PL04-00-0240"),
    ("Raskhody na piteyvuyu vodu",             "소모품비(음용수)",         "PL04-00-0090"),
    ("Raskhody na uslugi banka",               "지급수수료(은행수수료)",  "PL04-00-0240"),
    ("Raskhody na kantselyarskie prinadlezhnosti", "소모품비(문구)",      "PL04-00-0090"),
    ("Raskhody na khoz dlya sobstv nuzhd",     "소모품비(청소·잡비)",    "PL04-00-0090"),
    ("Registratsii onlayn KKM",                "지급수수료(금전등록기)",  "PL04-00-0240"),
    ("Remont org tekhniki ofisnoj",            "지급수수료(사무기기수리)","PL04-00-0240"),
    ("Sotsialnyy nalog",                       "복리후생비(사회보험세)",  "PL04-00-0050"),
    ("Strakhovoy polis",                       "보험료",                  "PL04-00-0150"),
    ("Sotovaya svyaz",                         "통신비(휴대폰)",          "PL04-00-0120"),
    ("Tekhnicheskoe obsluzhivanie KKM",        "지급수수료(금전등록기유지)", "PL04-00-0240"),
    ("Transportnye raskhody",                  "운반비(운송)",             "PL04-00-0110"),
    ("Uslugi po organizatsii dokumentooborota (Didox)", "지급수수료(전자문서·Didox)", "PL04-00-0240"),
    ("Uslugi po skladirovaniyu i khraneniyu gruza", "지급수수료(창고보관)", "PL04-00-0240"),
    ("Khranenie Informatsionnoy bazy v oblachnom servere", "지급수수료(클라우드)", "PL04-00-0240"),
    ("Informatsionno-tekhnicheskoe soprovozhdenie 1C 8.3", "지급수수료(1C시스템유지)", "PL04-00-0240"),
]

# ── 3. 엑셀 작성 ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "UZ01_계정마스터"

THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

HEADERS = [
    # (field,          표시명,                  fill_hex)
    ("subsidiary_code","법인코드",              "FFF2CC"),
    ("local_code",     "현지회계코드(1C)",      "FFF2CC"),
    ("local_name",     "현지회계 계정명",       "FFF2CC"),
    ("netra_code",     "네트라 계정코드",       "BDD7EE"),
    ("netra_name",     "네트라 계정명",         "BDD7EE"),
    ("confinas_code",  "Confinas 코드",         "D9EAD3"),
    ("confinas_name",  "Confinas 계정명",       "D9EAD3"),
    ("standard_code",  "신계정코드(FP/PL)",     "EFEFEF"),
    ("standard_name",  "신계정명(참고)",        "EFEFEF"),
    ("account_type",   "계정유형",              "EFEFEF"),
]

# 1행: 타이틀
ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
ws["A1"] = "UZ01 계정과목 마스터 매핑 양식  [황색=현지회계(입력됨)  청색=네트라 입력필요  녹색=Confinas 입력필요  회색=참고용]"
ws["A1"].font = Font(bold=True, size=12)
ws["A1"].fill = PatternFill("solid", fgColor="4472C4")
ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 22

# 2행: 헤더
for ci, (_, label, color) in enumerate(HEADERS, 1):
    c = ws.cell(row=2, column=ci, value=label)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=color)
    c.border = BORDER
    c.alignment = CENTER
ws.row_dimensions[2].height = 20

# 3행~: 데이터
for ri, (lc, ln, sc) in enumerate(UZ01_ROWS, start=3):
    std = std_lookup.get(sc, {"name": "", "type": ""})
    values = ["UZ01", lc, ln, "", "", "", "", sc, std["name"], std["type"]]
    fills  = ["FFF2CC","FFF2CC","FFF2CC","EBF3FB","EBF3FB","EAF1DD","EAF1DD","F2F2F2","F2F2F2","F2F2F2"]
    for ci, (val, fill) in enumerate(zip(values, fills), 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=False)

# 열 너비
widths = [10, 18, 28, 18, 22, 16, 20, 22, 24, 12]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{len(UZ01_ROWS)+2}"

# ── 안내 시트 ────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("작성안내")
guide = [
    ["항목",          "색상",    "설명"],
    ["subsidiary_code","황색",   "법인코드 — UZ01 고정값"],
    ["local_code",     "황색",   "1C 우즈베키스탄 계정코드 (숫자코드 또는 러시아어 비용항목명)"],
    ["local_name",     "황색",   "현지회계 계정 한국어명 (자동입력됨)"],
    ["netra_code",     "청색 ★", "네트라 계정코드 — 직접 입력 필요"],
    ["netra_name",     "청색 ★", "네트라 계정명 — 직접 입력 필요"],
    ["confinas_code",  "녹색 ★", "Confinas 업로드 계정코드 — 직접 입력 필요"],
    ["confinas_name",  "녹색 ★", "Confinas 계정명 — 직접 입력 필요"],
    ["standard_code",  "회색",   "내부 신계정코드(FP/PL) — 참고용, 수정불필요"],
    ["standard_name",  "회색",   "신계정 한국어명 — 참고용"],
    ["account_type",   "회색",   "계정유형 (asset/liability/equity/revenue/expense)"],
    ["",              "",        ""],
    ["★ N:1 매핑 허용", "",      "여러 현지(1C) 계정 → 하나의 네트라/Confinas 계정에 매핑 가능"],
    ["예시", "",                 "5010(현금)+5050(소매현금)+5110(당좌) 모두 → 네트라 101000(현금및현금성자산)"],
    ["", "",                     ""],
    ["★ 러시아어 항목", "",      "하단 PL 비용항목은 local_code 가 러시아어(발음 표기)로 되어있음"],
    ["주의", "",                 "현지회계 엑셀의 실제 항목명과 정확히 일치해야 매핑됩니다"],
]
for ri, row in enumerate(guide, 1):
    for ci, val in enumerate(row, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        if ri == 1:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="4472C4")
            c.font = Font(bold=True, color="FFFFFF")
        elif ci == 1 and ri > 1 and val:
            c.font = Font(bold=True)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 70

# ── 저장 ──────────────────────────────────────────────────────────────────────
out_dir = ROOT / "data/reference"
out_dir.mkdir(parents=True, exist_ok=True)
out_path = out_dir / "UZ01_계정마스터_양식.xlsx"
wb.save(str(out_path))
print("OK: " + str(out_path))
print("count: " + str(len(UZ01_ROWS)))
