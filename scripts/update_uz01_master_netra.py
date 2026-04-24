# -*- coding: utf-8 -*-
"""기존 UZ01_계정마스터_양식.xlsx 에서
  - 네트라 계정코드 / 네트라 계정명  열 제거
  - 네트라 항목 열 추가 (5개 dropdown)
  - Confinas 데이터 보존

실행: python scripts/update_uz01_master_netra.py
"""
import pathlib
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT    = pathlib.Path(__file__).resolve().parent.parent
SRC     = ROOT / "data/reference/UZ01_계정마스터_양식.xlsx"
OUT     = ROOT / "data/reference/UZ01_계정마스터_양식.xlsx"

NETRA_CATS = ["매출채권", "선수금", "원가", "재고자산", "매출액"]
STEP1_CATS = [
    "(HQ) PRODUCT",
    "(KR) MERCHANDISE",
    "(US) PRODUCT",
    "(Relative) PRODUCT",
    "(Domestic) MERCHANDISE",
]

THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── 새 컬럼 정의 ─────────────────────────────────────────────────────────────
# (field_key, 헤더명, fill_hex, width)
NEW_HEADERS = [
    ("subsidiary_code", "법인코드",              "FFF2CC", 10),
    ("local_code",      "현지회계코드(1C)",       "FFF2CC", 18),
    ("local_name",      "현지회계 계정명",        "FFF2CC", 28),
    ("netra_category",  "네트라 항목",            "D0E4FF", 14),
    ("netra_step1",     "네트라 Step1",           "D0E4FF", 22),  # 청색 동일
    ("confinas_code",   "Confinas 코드",         "D9EAD3", 16),
    ("confinas_name",   "Confinas 계정명",       "D9EAD3", 20),
    ("standard_code",   "신계정코드(FP/PL)",      "EFEFEF", 22),
    ("standard_name",   "신계정명(참고)",         "EFEFEF", 24),
    ("account_type",    "계정유형",               "EFEFEF", 12),
]

# 기존 헤더 → field_key 매핑 (읽기용)
OLD_HEADER_MAP = {
    "법인코드":           "subsidiary_code",
    "현지회계코드(1c)":   "local_code",
    "현지회계 계정명":    "local_name",
    "네트라 계정코드":    "netra_code_OLD",      # 삭제 대상
    "네트라 step1":      "netra_step1",
    "네트라 Step1":      "netra_step1",
    "네트라 계정명":      "netra_name_OLD",      # 삭제 대상
    "confinas 코드":     "confinas_code",
    "confinas 계정명":   "confinas_name",
    "신계정코드(fp/pl)":  "standard_code",
    "신계정명(참고)":     "standard_name",
    "계정유형":           "account_type",
}

wb_src = openpyxl.load_workbook(SRC)
ws_src = wb_src["UZ01_계정마스터"]

# 기존 헤더행(2행) 파악
header_row = 2
old_col_map = {}   # col_idx → field_key
for ci in range(1, ws_src.max_column + 1):
    val = ws_src.cell(header_row, ci).value
    if val:
        key = OLD_HEADER_MAP.get(str(val).strip().lower(), None)
        if key:
            old_col_map[ci] = key

# 기존 데이터 읽기 (3행~)
old_data = []
for ri in range(3, ws_src.max_row + 1):
    row = {}
    for ci, field in old_col_map.items():
        row[field] = ws_src.cell(ri, ci).value
    if not row.get("subsidiary_code"):
        continue
    old_data.append(row)

wb_src.close()

# ── 새 워크북 작성 ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "UZ01_계정마스터"

# 1행: 타이틀
ws.merge_cells(f"A1:{get_column_letter(len(NEW_HEADERS))}1")
ws["A1"] = (
    "UZ01 계정과목 마스터  "
    "[황색=현지회계(입력됨)  청색=네트라항목(입력필요: 매출채권/선수금/원가/재고자산/매출액)  "
    "녹색=Confinas(입력됨)  회색=참고용]"
)
ws["A1"].fill = PatternFill("solid", fgColor="4472C4")
ws["A1"].font = Font(bold=True, size=11, color="FFFFFF")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 28

# 2행: 헤더
for ci, (_, label, color, _w) in enumerate(NEW_HEADERS, 1):
    c = ws.cell(2, ci, label)
    c.font   = Font(bold=True, size=10)
    c.fill   = PatternFill("solid", fgColor=color)
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# 3행~: 데이터
for ri, d in enumerate(old_data, start=3):
    for ci, (field, _, color, _w) in enumerate(NEW_HEADERS, 1):
        val = d.get(field, "")
        c = ws.cell(ri, ci, val if val else "")
        c.fill   = PatternFill("solid", fgColor=color if color != "D0E4FF" else "EBF3FB")
        c.border = BORDER
        c.alignment = Alignment(vertical="center")

# Dropdown: netra_category
netra_col_idx = next(
    ci for ci, (f, *_) in enumerate(NEW_HEADERS, 1) if f == "netra_category"
)
dv = DataValidation(
    type="list",
    formula1='"' + ",".join(NETRA_CATS) + '"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="입력오류",
    error="매출채권 / 선수금 / 원가 / 재고자산 / 매출액 중 선택하세요.",
)
dv.sqref = (
    f"{get_column_letter(netra_col_idx)}3:"
    f"{get_column_letter(netra_col_idx)}{len(old_data)+10}"
)
ws.add_data_validation(dv)

# Dropdown: netra_step1
step1_col_idx = next(
    ci for ci, (f, *_) in enumerate(NEW_HEADERS, 1) if f == "netra_step1"
)
dv2 = DataValidation(
    type="list",
    formula1='"' + ",".join(STEP1_CATS) + '"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="입력오류",
    error="(HQ) PRODUCT / (KR) MERCHANDISE / (US) PRODUCT / (Relative) PRODUCT / (Domestic) MERCHANDISE 중 선택하세요.",
)
dv2.sqref = (
    f"{get_column_letter(step1_col_idx)}3:"
    f"{get_column_letter(step1_col_idx)}{len(old_data)+10}"
)
ws.add_data_validation(dv2)

# 열 너비
for ci, (_, _, _, w) in enumerate(NEW_HEADERS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(NEW_HEADERS))}{len(old_data)+2}"

# 안내 시트 업데이트
if "작성안내" in wb.sheetnames:
    del wb["작성안내"]
ws2 = wb.create_sheet("작성안내")
guide = [
    ["항목",             "색상",    "설명"],
    ["subsidiary_code", "황색",    "법인코드 — UZ01 고정"],
    ["local_code",      "황색",    "1C 현지회계 계정코드"],
    ["local_name",      "황색",    "현지회계 계정 한국어명"],
    ["netra_category",  "청색 ★", "네트라 대사 항목 — 드롭다운 선택 (아래 5개 중)"],
    ["",                "",        "  · 매출채권  · 선수금  · 원가  · 재고자산  · 매출액"],
    ["",                "",        "  해당 없으면 빈칸"],
    ["netra_step1",     "청색 ★", "네트라 Step1 구분 — 드롭다운 선택 (아래 5개 중)"],
    ["",                "",        "  · (HQ) PRODUCT  · (KR) MERCHANDISE  · (US) PRODUCT"],
    ["",                "",        "  · (Relative) PRODUCT  · (Domestic) MERCHANDISE"],
    ["",                "",        "  매출액/원가/재고자산 계정만 입력, 매출채권/선수금은 빈칸"],
    ["confinas_code",   "녹색",    "Confinas 업로드 코드 (이미 입력됨)"],
    ["confinas_name",   "녹색",    "Confinas 계정명 (이미 입력됨)"],
    ["standard_code",   "회색",    "내부 신계정코드 — 참고용"],
    ["standard_name",   "회색",    "신계정명 — 참고용"],
    ["account_type",    "회색",    "계정유형 — 참고용"],
    ["", "", ""],
    ["★ N:1 매핑", "", "여러 현지 계정이 같은 네트라 항목에 속할 수 있음"],
    ["예시", "", "5010(현금)+5110(당좌)+5210(외화) 모두 → 매출채권 아니라 해당없음 (네트라 비교 대상 아님)"],
    ["예시", "", "4010(매출채권)+4015(외화매출채권) → 네트라 항목: 매출채권"],
]
for ri, row in enumerate(guide, 1):
    for ci, val in enumerate(row, 1):
        c = ws2.cell(ri, ci, val)
        if ri == 1:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="4472C4")
            c.font = Font(bold=True, color="FFFFFF")
        elif ci == 1 and val:
            c.font = Font(bold=True)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 65

wb.save(str(OUT))
print("OK: " + str(OUT))
print("rows: " + str(len(old_data)))
print("netra_col: " + str(netra_col_idx))
