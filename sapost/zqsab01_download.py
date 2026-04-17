"""
ZQSAB01 — QSA: 매출마감 실적현황 조회 (품목별 연결손익) 자동화 모듈

선택 화면 구조 (diagnose_zqsab01.py 확인 결과):
  - 품목코드 : wnd[0]/usr/ctxtP_PCODE  (선택 필터)
  - 작업기간 : wnd[0]/usr/txtP_PERIO   (기간 입력)
  - [결과 리포트 Excel] 버튼 → Excel 파일 직접 다운로드
  - [인터랙트 조회]    버튼 → ALV 결과 화면

흐름:
  connect() → navigate() → set_params_and_execute()
  → wait_for_download() → process_excel() → close()

실행 예 (직접):
  python sapost/zqsab01_download.py --perio 202603
  python sapost/zqsab01_download.py --perio 202603 --pcode KR001
"""

import sys
import time
import shutil
import logging
import argparse
import configparser
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sapost.src.utils import get_config, setup_logger


# ─────────────────────────────────────────────────────────────
# ZQSAB01 Downloader
# ─────────────────────────────────────────────────────────────

class ZQSAB01Downloader:
    """
    ZQSAB01 품목별 연결손익 조회 자동화 클래스.

    '결과 리포트 Excel' 버튼을 클릭해 SAP에서 Excel 파일을 직접 다운로드하고
    output_dir에 정리·저장합니다.
    """

    TRANSACTION = "ZQSAB01"

    def __init__(self, config: configparser.ConfigParser, logger: logging.Logger):
        if not WIN32COM_AVAILABLE:
            raise ImportError(
                "pywin32가 설치되어 있지 않습니다.\n"
                "pip install pywin32  을 실행하세요."
            )
        self.config  = config
        self.logger  = logger
        self.session = None

        sec = "ZQSAB01"
        self.pcode_field  = config.get(sec, "pcode_field")
        self.perio_field  = config.get(sec, "perio_field")
        self.excel_btn    = config.get(sec, "excel_btn")
        self.download_dir = Path(config.get(sec, "download_dir",
                                            fallback="C:/Users/Osstem/Downloads"))
        self.output_dir   = Path(config.get(sec, "output_dir",
                                            fallback="sapost/data/output/zqsab01"))
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ── 공개 메서드 ──────────────────────────────────────────

    def connect(self):
        """실행 중인 SAP GUI 세션에 연결합니다."""
        self.logger.info("SAP GUI 세션 연결 중...")
        try:
            sap_gui_auto = win32com.client.GetObject("SAPGUI")
            application  = sap_gui_auto.GetScriptingEngine
            connection   = application.Children(0)
            self.session = connection.Children(0)
            self.logger.info("SAP GUI 세션 연결 완료")
        except Exception as e:
            raise ConnectionError(
                f"SAP GUI 세션에 연결할 수 없습니다: {e}\n"
                "SAP GUI가 실행 중이고 스크립팅이 활성화되어 있는지 확인하세요."
            )

    def navigate(self):
        """ZQSAB01 트랜잭션으로 이동합니다."""
        self.logger.info(f"트랜잭션 이동: {self.TRANSACTION}")
        try:
            cmd = self.session.findById("wnd[0]/tbar[0]/okcd")
            cmd.text = f"/n{self.TRANSACTION}"
            self.session.findById("wnd[0]").sendVKey(0)  # Enter
            time.sleep(2)
            self.logger.info("트랜잭션 이동 완료")
        except Exception as e:
            raise RuntimeError(f"트랜잭션 이동 실패: {e}")

    def set_params_and_execute(self, perio: str, pcode: str = ""):
        """
        선택 화면에 기간·품목코드를 입력하고 '결과 리포트 Excel' 버튼을 클릭합니다.

        Args:
            perio: 작업기간 (예: '202603' — YYYYMM 또는 SAP 기간 형식)
            pcode: 품목코드 필터 (빈칸이면 전체 조회)
        """
        self.logger.info(
            f"파라미터 입력 — 작업기간:{perio}"
            + (f"  품목코드:{pcode}" if pcode else "  (품목코드 전체)")
        )
        try:
            # 품목코드 입력 (선택)
            if pcode:
                self.session.findById(self.pcode_field).text = pcode

            # 작업기간 입력
            self.session.findById(self.perio_field).text = perio

            # '결과 리포트 Excel' 버튼 클릭
            self.logger.info("'결과 리포트 Excel' 버튼 클릭")
            self.session.findById(self.excel_btn).press()
            time.sleep(2)

            # 다운로드 대화상자 처리 (열리는 경우)
            self._handle_save_dialog(perio)

        except Exception as e:
            raise RuntimeError(f"파라미터 입력/실행 실패: {e}")

    def wait_for_download(self, perio: str, timeout: float = 60.0) -> Path:
        """
        다운로드 폴더에서 새로 생성된 Excel 파일을 기다렸다가 output_dir로 이동합니다.

        반환: output_dir에 저장된 파일 경로
        """
        self.logger.info(f"다운로드 대기 중 (최대 {timeout}초)...")
        deadline = time.time() + timeout
        before   = set(self.download_dir.glob("*.xlsx")) | set(self.download_dir.glob("*.xls"))

        while time.time() < deadline:
            time.sleep(1)
            after = set(self.download_dir.glob("*.xlsx")) | set(self.download_dir.glob("*.xls"))
            new_files = after - before
            if new_files:
                # 가장 최근 파일 선택
                downloaded = max(new_files, key=lambda p: p.stat().st_mtime)
                dest = self.output_dir / f"ZQSAB01_{perio}{downloaded.suffix}"
                shutil.move(str(downloaded), str(dest))
                self.logger.info(f"파일 이동 완료: {dest}")
                return dest

        raise TimeoutError(
            f"다운로드 파일을 {timeout}초 내에 찾지 못했습니다.\n"
            f"다운로드 폴더: {self.download_dir}"
        )

    def process_excel(self, file_path: Path, perio: str) -> Path:
        """
        SAP에서 다운로드된 Excel을 읽어 서식·헤더를 정리한 결과 파일로 재저장합니다.

        반환: 최종 결과 파일 경로
        """
        self.logger.info(f"Excel 가공 시작: {file_path}")
        try:
            df = pd.read_excel(file_path, header=0)
        except Exception as e:
            self.logger.warning(f"Excel 읽기 실패 ({e}) — 원본 파일을 그대로 사용합니다.")
            return file_path

        # 빈 행·열 제거
        df = df.dropna(how="all").reset_index(drop=True)
        df = df.loc[:, df.notna().any()]

        out_path = self.output_dir / f"ZQSAB01_{perio}_processed.xlsx"
        _write_styled_excel(df, out_path, title=f"품목별 연결손익  [{perio}]")
        self.logger.info(f"가공 완료: {out_path}")
        return out_path

    def close(self):
        """SAP 초기 화면으로 복귀합니다."""
        try:
            if self.session:
                self.session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                self.session.findById("wnd[0]").sendVKey(0)
                self.logger.info("SAP 초기 화면 복귀")
        except Exception:
            pass

    # ── 내부 메서드 ─────────────────────────────────────────

    def _handle_save_dialog(self, perio: str):
        """
        SAP '다른 이름으로 저장' 또는 파일명 입력 대화상자가 열리면 처리합니다.
        열리지 않으면 자동으로 다운로드 폴더에 저장됩니다.
        """
        time.sleep(1)
        try:
            # 대화상자 wnd[1] 존재 여부 확인
            dialog = self.session.findById("wnd[1]")
            self.logger.info("파일 저장 대화상자 감지")

            # 파일명 입력 필드 찾기 (SAP 버전에 따라 다를 수 있음)
            try:
                filename_field = dialog.findById("usr/ctxtDY_FILENAME")
                filename_field.text = f"ZQSAB01_{perio}.xlsx"
            except Exception:
                pass

            # '교체(저장)' 또는 '확인' 버튼
            try:
                dialog.findById("tbar[0]/btn[11]").press()  # 교체
            except Exception:
                try:
                    dialog.findById("tbar[0]/btn[0]").press()  # 확인
                except Exception:
                    pass
            time.sleep(2)
        except Exception:
            # 대화상자 없음 — 브라우저 다운로드 방식
            pass


# ─────────────────────────────────────────────────────────────
# Excel 서식 출력 헬퍼
# ─────────────────────────────────────────────────────────────

def _write_styled_excel(df: pd.DataFrame, out_path: Path, title: str = ""):
    """DataFrame을 서식 있는 Excel로 저장합니다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "품목별 연결손익"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    row_offset = 0

    # 제목 행
    if title:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=max(len(df.columns), 1))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font      = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        row_offset = 1

    # 헤더
    hdr_row = row_offset + 1
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=hdr_row, column=c, value=col)
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin
    ws.row_dimensions[hdr_row].height = 28

    # 데이터
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    for r, row in enumerate(df.itertuples(index=False), start=hdr_row + 1):
        for c, val in enumerate(row, 1):
            v    = _coerce(val)
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin
            if isinstance(v, (int, float)):
                cell.alignment    = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.##"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if r % 2 == 0:
                cell.fill = alt_fill

    # 열 너비
    for c, col in enumerate(df.columns, 1):
        vals   = [str(col)] + [str(v) for v in df.iloc[:, c - 1] if pd.notna(v)]
        max_w  = max((len(v) for v in vals), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(max_w + 2, 40)

    ws.freeze_panes = f"A{hdr_row + 1}"
    wb.save(out_path)


def _coerce(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    cleaned = s.replace(".", "").replace(",", ".")
    try:
        return float(cleaned) if "." in cleaned else int(cleaned)
    except ValueError:
        return s


# ─────────────────────────────────────────────────────────────
# CLI 직접 실행
# ─────────────────────────────────────────────────────────────

def _parse_args():
    parser = argparse.ArgumentParser(
        description="ZQSAB01 품목별 연결손익 조회 자동화"
    )
    parser.add_argument(
        "--perio",
        default=datetime.now().strftime("%Y%m"),
        help="작업기간 (예: 202603). 기본값: 이번 달",
    )
    parser.add_argument(
        "--pcode",
        default="",
        help="품목코드 필터 (생략 시 전체 조회)",
    )
    return parser.parse_args()


def main():
    args   = _parse_args()
    config = get_config()
    logger = setup_logger("sapost.zqsab01", config)

    logger.info("=" * 60)
    logger.info(f"ZQSAB01 품목별 연결손익 조회 시작 | 기간:{args.perio}"
                + (f"  품목:{args.pcode}" if args.pcode else ""))
    logger.info("=" * 60)

    dl = ZQSAB01Downloader(config, logger)
    dl.connect()
    try:
        dl.navigate()
        dl.set_params_and_execute(perio=args.perio, pcode=args.pcode)
        raw  = dl.wait_for_download(args.perio)
        out  = dl.process_excel(raw, args.perio)
        logger.info(f"완료. 결과 파일: {out}")
    finally:
        dl.close()


if __name__ == "__main__":
    main()
