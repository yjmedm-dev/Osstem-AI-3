"""
양식 붙여넣기 모듈
정제된 DataFrame을 고정 템플릿 엑셀 파일의 지정 위치에 기입
"""
import shutil
import logging
import configparser
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
from pathlib import Path


class TemplateWriter:
    def __init__(self, config: configparser.ConfigParser, logger: logging.Logger):
        self.config = config
        self.logger = logger

        self.template_path = Path(config.get("PATHS", "template_path"))
        self.output_dir = Path(config.get("PATHS", "output_dir"))
        self.sheet_name = config.get("TEMPLATE", "sheet_name", fallback="Sheet1")
        self.start_cell = config.get("TEMPLATE", "start_cell", fallback="A2")

        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write(self, df: pd.DataFrame, month: str) -> Path:
        """
        df: 정제된 데이터
        month: 기준월 (예: "202503")
        반환: 저장된 결과 파일 경로
        """
        if not self.template_path.exists():
            raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {self.template_path}")

        output_path = self.output_dir / f"SAP_{month}.xlsx"
        shutil.copy2(self.template_path, output_path)
        self.logger.info(f"템플릿 복사 완료: {output_path}")

        wb = openpyxl.load_workbook(output_path)

        if self.sheet_name not in wb.sheetnames:
            raise ValueError(
                f"시트 '{self.sheet_name}'가 없습니다. "
                f"존재하는 시트: {wb.sheetnames}"
            )

        ws = wb[self.sheet_name]

        start_col, start_row = self._parse_cell(self.start_cell)

        self._clear_data_area(ws, start_row, start_col, df)
        self.logger.info("기존 데이터 영역 초기화 완료")

        for r_offset, row_data in enumerate(df.itertuples(index=False)):
            for c_offset, value in enumerate(row_data):
                cell = ws.cell(
                    row=start_row + r_offset,
                    column=start_col + c_offset,
                )
                cell.value = None if pd.isna(value) else value

        wb.save(output_path)
        self.logger.info(f"양식 저장 완료: {output_path} ({len(df)}행 기입)")

        return output_path

    def _parse_cell(self, cell_str: str) -> tuple[int, int]:
        """'B3' → (col=2, row=3) 반환"""
        cell_str = cell_str.strip().upper()
        col_str = "".join(c for c in cell_str if c.isalpha())
        row_str = "".join(c for c in cell_str if c.isdigit())
        return column_index_from_string(col_str), int(row_str)

    def _clear_data_area(self, ws, start_row: int, start_col: int, df: pd.DataFrame):
        """기존 데이터가 있을 범위를 빈 값으로 초기화"""
        num_rows = ws.max_row - start_row + 1
        num_cols = len(df.columns)
        for r in range(start_row, start_row + max(num_rows, 1)):
            for c in range(start_col, start_col + num_cols):
                ws.cell(row=r, column=c).value = None
